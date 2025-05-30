import io
import boto3
from botocore.exceptions import NoCredentialsError
import openpyxl
from openpyxl.utils import get_column_letter
import os

def upload_image_to_s3(image_bytes_io, s3_path):
    try:
        s3 = boto3.client(
            "s3",
            aws_access_key_id=st.secrets["aws"]["aws_access_key_id"],
            aws_secret_access_key=st.secrets["aws"]["aws_secret_access_key"],
            region_name=st.secrets["aws"]["aws_region"]
        )
        s3.upload_fileobj(image_bytes_io, st.secrets["aws"]["bucket_name"], s3_path)
        return True, f"Uploaded {s3_path}"
    except NoCredentialsError:
        return False, "AWS credentials not found"
    except Exception as e:
        return False, str(e)


def append_student_to_excel(enrollment, name, excel_path="students.xlsx"):
    """
    Append a student entry to an Excel file.
    If file doesn't exist, create it with header.

    Args:
        enrollment (str): Enrollment number
        name (str): Student name
        excel_path (str): Path to Excel file (local path)

    Returns:
        None
    """
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Create headers
        ws.append(["Enrollment", "Name"])

    # Append the new student
    ws.append([enrollment, name])

    wb.save(excel_path)


def add_student_with_augmented_images_to_s3(enrollment, name, uploaded_image):
    """
    Generate 100 augmented images from uploaded_image and upload all to S3 under
    student_images/{class_name}/{enrollment}_{name}_{index}.jpg
    Also append student data to an Excel file.

    Args:
        enrollment (str): Enrollment number
        name (str): Student name (spaces replaced with underscores)
        uploaded_image (UploadedFile): Streamlit uploaded image file

    Returns:
        (bool, str): success flag and message
    """
    try:
        import streamlit as st

        clean_name = name.strip().replace(" ", "_")
        student_class = "default_class"  # Replace if needed

        # Read image from bytes (OpenCV format)
        img = read_image_from_bytes(uploaded_image)
        if img is None:
            return False, "Failed to read uploaded image."

        # Generate 100 augmented images
        augmented_images = generate_augmented_images(img, count=100)

        failed_uploads = []
        for i, aug_img in enumerate(augmented_images):
            # Convert OpenCV image to JPEG bytes
            is_success, buffer = cv2.imencode(".jpg", aug_img)
            if not is_success:
                failed_uploads.append(f"{enrollment}_{clean_name}_{i+1}.jpg (encode fail)")
                continue

            img_bytes = io.BytesIO(buffer.tobytes())
            s3_path = f"student_images/{student_class}/{enrollment}_{clean_name}_{i+1}.jpg"

            success, msg = upload_image_to_s3(img_bytes, s3_path)
            if not success:
                failed_uploads.append(f"{enrollment}_{clean_name}_{i+1}.jpg ({msg})")

        if failed_uploads:
            return False, f"Some images failed to upload: {failed_uploads}"

        # Append student info to Excel (local file here)
        append_student_to_excel(enrollment, name)

        return True, f"All 100 augmented images uploaded successfully for student {name}."

    except Exception as e:
        return False, f"Error: {str(e)}"
